"""
RevOS Data Anonymizer
Reads the 'Raw Input' tab from the template, assigns Customer 1/2/3... IDs,
generates the 'Customer Key' mapping tab, and populates the 'Deal Data' tab.

Usage:
    python revos-anonymize.py <path-to-xlsx>
    python revos-anonymize.py RevOS-Data-Template.xlsx
"""

import sys
from pathlib import Path
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def anonymize(filepath):
    filepath = Path(filepath)
    if not filepath.exists():
        print(f"Error: {filepath} not found")
        sys.exit(1)

    print(f"Loading {filepath}...")
    wb = load_workbook(filepath)

    # ─── Read Raw Input ───
    if "Raw Input" not in wb.sheetnames:
        print("Error: 'Raw Input' tab not found. Paste your data there first.")
        sys.exit(1)

    ws_raw = wb["Raw Input"]
    # Find data start (row 5, after header + notes)
    data_start = 5
    rows = []
    for r in range(data_start, ws_raw.max_row + 1):
        account = ws_raw.cell(row=r, column=1).value
        if not account or str(account).strip() == "":
            continue
        row_data = []
        for c in range(1, 13):  # 12 columns
            row_data.append(ws_raw.cell(row=r, column=c).value)
        rows.append(row_data)

    if not rows:
        print("Error: No data found in 'Raw Input'. Paste your deals starting at row 5.")
        sys.exit(1)

    print(f"Found {len(rows)} deal records.")

    # ─── Build Customer Key ───
    # Preserve order of first appearance
    customer_map = OrderedDict()
    customer_counts = {}
    customer_verticals = {}

    for row in rows:
        name = str(row[0]).strip()
        if name not in customer_map:
            idx = len(customer_map) + 1
            customer_map[name] = f"Customer {idx}"
            customer_counts[name] = 0
            customer_verticals[name] = row[11] if row[11] else ""
        customer_counts[name] += 1

    print(f"Found {len(customer_map)} unique customers:")
    for real_name, anon_name in customer_map.items():
        print(f"  {anon_name:15s} ← {real_name} ({customer_counts[real_name]} deals)")

    # ─── Populate Customer Key tab ───
    if "Customer Key" not in wb.sheetnames:
        ws_key = wb.create_sheet("Customer Key", 1)
    else:
        ws_key = wb["Customer Key"]

    # Clear existing data (keep header rows 1-3)
    for r in range(4, ws_key.max_row + 1):
        for c in range(1, 5):
            ws_key.cell(row=r, column=c).value = None

    # Styles
    thin = Side(style="thin", color="D0D0D0")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    body_font = Font(name="Arial", size=10)
    purple_font = Font(name="Arial", bold=True, size=10, color="7B2CBF")
    key_fill = PatternFill("solid", fgColor="E8DAEF")
    name_fill = PatternFill("solid", fgColor="FEF3C7")

    for i, (real_name, anon_name) in enumerate(customer_map.items()):
        r = 4 + i
        # Anonymized name
        cell = ws_key.cell(row=r, column=1, value=anon_name)
        cell.font = purple_font
        cell.fill = key_fill
        cell.border = border
        # Real name
        cell = ws_key.cell(row=r, column=2, value=real_name)
        cell.font = body_font
        cell.fill = name_fill
        cell.border = border
        # Vertical
        cell = ws_key.cell(row=r, column=3, value=customer_verticals[real_name])
        cell.font = body_font
        cell.border = border
        # Deal count
        cell = ws_key.cell(row=r, column=4, value=customer_counts[real_name])
        cell.font = body_font
        cell.border = border

    # ─── Populate Deal Data tab ───
    if "Deal Data" not in wb.sheetnames:
        ws_deal = wb.create_sheet("Deal Data", 2)
    else:
        ws_deal = wb["Deal Data"]

    # Clear existing data (keep header rows 1-3)
    for r in range(4, ws_deal.max_row + 1):
        for c in range(1, 13):
            ws_deal.cell(row=r, column=c).value = None

    for i, row_data in enumerate(rows):
        r = 4 + i
        real_name = str(row_data[0]).strip()
        anon_name = customer_map.get(real_name, real_name)

        for c in range(1, 13):
            val = row_data[c - 1] if c <= len(row_data) else None
            # Replace column 1 (Customer Account) with anonymized name
            if c == 1:
                val = anon_name

            cell = ws_deal.cell(row=r, column=c, value=val)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="top")

            if c == 1:
                cell.font = Font(name="Arial", bold=True, size=10, color="7B2CBF")
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
            elif isinstance(val, (int,)) and c == 10:
                cell.number_format = '#,##0'

    # ─── Save ───
    output = filepath
    wb.save(output)
    print(f"\nSaved to {output}")
    print(f"  'Customer Key' tab: {len(customer_map)} customers mapped")
    print(f"  'Deal Data' tab: {len(rows)} anonymized deal records")
    print(f"\n  Export 'Deal Data' tab as CSV to upload to RevOS.")
    print(f"  Keep 'Customer Key' tab as your private reference.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python revos-anonymize.py <path-to-xlsx>")
        print("Example: python revos-anonymize.py RevOS-Data-Template.xlsx")
        sys.exit(1)
    anonymize(sys.argv[1])
